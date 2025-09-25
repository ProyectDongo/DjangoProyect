from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from .models import ExerciseLog, TrainingPlan, Workout, WorkoutExercise, Warmup,Exercise
from .forms import TrainingPlanForm, WorkoutForm, WorkoutExerciseForm, ExerciseLogForm, WarmupForm
from core.models import User
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from .forms import ClientCreationForm,ExerciseForm
from django.utils.crypto import get_random_string
import mimetypes
from io import BytesIO
import openpyxl
import json
from django.core.mail import EmailMessage
from django.db.models import Avg
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import boto3
from botocore.exceptions import ClientError
import uuid 
from django.core.validators import FileExtensionValidator

# ====================================================================================================================
# Introducción al Archivo de Vistas
# ====================================================================================================================
# Este archivo contiene todas las vistas (views) relacionadas con la funcionalidad principal de la aplicación de entrenamiento físico.
# Como creador de este código, lo diseñé con un enfoque en la modularidad, seguridad y eficiencia. Cada vista está decorada con @login_required
# para asegurar que solo usuarios autenticados puedan acceder, y verifico roles explícitamente para prevenir accesos no autorizados.
#
# Las vistas se dividen en secciones temáticas: dashboards para entrenadores y clientes, creación y gestión de clientes, planes, workouts,
# ejercicios, calentamientos y logs. Utilizo Querysets de Django para consultas eficientes, evitando problemas como N+1 queries donde sea posible
# (por ejemplo, usando .filter() y .count() en lugar de loops innecesarios). Los forms se usan para validación y sanitización de inputs del usuario.
#
# Por qué este enfoque: La app maneja datos sensibles (como progresos de clientes y emails), por lo que priorizo la seguridad (e.g., get_object_or_404
# para manejar 404s graciosamente y verificar ownership). También, calculo métricas en el backend para reducir carga en el frontend y mejorar la UX.
#
# Nota: Todas las redirecciones usan nombres de URLs reversibles (e.g., 'trainer_dashboard') para mantener la mantenibilidad si las URLs cambian.
# ====================================================================================================================


#----------------------------Vistas para el Dashboard del Entrenador - El Centro de Control---------------------------------

# Esta sección contiene vistas exclusivas para entrenadores. Comienzo con trainer_dashboard, que sirve como el "hub" principal donde los entrenadores
# pueden ver un overview de sus planes, clientes y métricas clave. Verifico el rol inmediatamente para redirigir a usuarios no autorizados, previniendo
# accesos indebidos y mejorando la seguridad.
#
# Por qué: Un dashboard centralizado facilita la gestión diaria. Calculo conteos como active_plans_count usando .filter().count() para eficiencia,
# ya que evita cargar objetos completos en memoria. Para clientes, agrego atributos dinámicos (active_plans y last_session) en un loop simple; esto es
# aceptable porque el número de clientes por entrenador es bajo (asumiendo <100). Si escala, consideraría prefetch_related para optimizar.
#
# Los calentamientos se dividen por tipo para una presentación más organizada en la plantilla, permitiendo al entrenador recomendarlos fácilmente.
# El contexto es un diccionario rico que pasa todos los datos necesarios a la plantilla 'entrenador/entrenador.html', promoviendo separación de concerns.
@login_required
def trainer_dashboard(request):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')

    plans = TrainingPlan.objects.filter(trainer=request.user)
    active_plans_count = plans.filter(status='active').count()

    clients = User.objects.filter(role='CLIENTE', assigned_professional=request.user)
    clients_count = clients.count()

    weekly_sessions = Workout.objects.filter(
        plan__in=plans,
        date__range=[timezone.now().date(), timezone.now().date() + timedelta(days=7)]
    ).exclude(date__isnull=True).count()

    exercises_count = WorkoutExercise.objects.filter(workout__plan__in=plans).count()

    warmups = Warmup.objects.all()
    warmups_upper_body = warmups.filter(type='superior')
    warmups_lower_body = warmups.filter(type='inferior')

    for client in clients:
        client.active_plans = client.assigned_plans.filter(status='active').count()
        last_log = ExerciseLog.objects.filter(client=client).order_by('-date_completed').first()
        client.last_session = last_log.date_completed if last_log else None
        client.active_plan = client.assigned_plans.filter(status='active').first()  # NUEVA LÍNEA: Calcula aquí el plan activo

    context = {
        'plans': plans,
        'active_plans_count': active_plans_count,
        'clients_count': clients_count,
        'weekly_sessions': weekly_sessions,
        'exercises_count': exercises_count,
        'clients': clients,
        'warmups': {
            'upper_body': warmups_upper_body,
            'lower_body': warmups_lower_body,
        }
    }
    return render(request, 'entrenador/entrenador.html', context)


# Vista para detalles de un workout específico. Similar al dashboard, pero enfocada en un workout. Verifico rol y ownership con get_object_or_404.
#
# Por qué: Proporciona una vista granular para inspeccionar/editar workouts. Ordenamos ejercicios por 'order' para mantener la secuencia lógica.
# Esto asegura que la UI muestre ejercicios en el orden planeado, mejorando la usabilidad.

@login_required
def workout_detail(request, workout_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    # get_object_or_404: Maneja 404 si no existe o no pertenece al trainer. Evita crashes y verifica ownership.
    workout = get_object_or_404(Workout, id=workout_id, plan__trainer=request.user)
    exercises = workout.exercises.all().order_by('order')  # Orden por 'order' para secuencia correcta.
    context = {
        'workout': workout,
        'exercises': exercises,
    }
    return render(request, 'entrenador/workout_detail.html', context)


#----------------------- Creación de Clientes - Facilitando la Onboarding --------------------------------------------

# Vista para crear clientes. Permitida solo para ENTRENADOR o NUTRICIONISTA. Usa ClientCreationForm para validación.
#
# Por qué: Facilita el onboarding de clientes nuevos. Generamos una contraseña temporal segura con get_random_string para seguridad inicial.
# Asignamos rol y assigned_professional automáticamente. Enviamos mensajes de éxito/error con django.contrib.messages para feedback UX.
# Redirigimos al dashboard tras éxito para flujo continuo.

@login_required
def create_client(request):
    if request.user.role not in ['ENTRENADOR', 'NUTRICIONISTA']:  # Restricción de roles: Solo profesionales pueden crear clientes.
        messages.error(request, "No tienes permiso para crear clientes.")
        return redirect('inicio')

    if request.method == 'POST':
        form = ClientCreationForm(request.POST)  # Form para validación de inputs.
        if form.is_valid():
            client = form.save(commit=False)  # No commiteamos aún para agregar campos custom.
            client.role = 'CLIENTE'  # Asignación automática de rol.
            client.assigned_professional = request.user  # Asignación al creador.

            # Generar contraseña temporal: Longitud 12 para balance entre seguridad y usabilidad.
            temp_password = get_random_string(length=12)
            client.set_password(temp_password)  # Hashing seguro con Django's auth.
            client.save()
            # Mensaje de éxito incluye contraseña temporal (enviar por email en producción para más seguridad).
            messages.success(request, f"Cliente {client.username} creado exitosamente. Contraseña temporal: {temp_password}")
            return redirect('trainer_dashboard')
        else:
            messages.error(request, "Error al crear el cliente. Revisa los datos ingresados.")
    else:
        form = ClientCreationForm()  # Form vacío para GET.

    return render(request, 'entrenador/crear_cliente.html', {'form': form})


#-------------------------------#Gestión de Planes - El Corazón de la App--------------------------------------------

# Vista para crear planes de entrenamiento. Similar a create_client, pero para planes. Limito queryset de clientes a los asignados.
#
# Por qué: Los planes son centrales; esta vista asegura que solo se asignen a clientes del profesional. Redirigimos a detail tras creación para edición inmediata.

@login_required
def create_plan(request):
    if request.user.role not in ['ENTRENADOR', 'NUTRICIONISTA']:
        messages.error(request, "No tienes permiso para crear planes.")
        return redirect('inicio')

    if request.method == 'POST':
        form = TrainingPlanForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.trainer = request.user  # Asignación automática.
            plan.save()
            messages.success(request, "Plan creado exitosamente.")
            return redirect('trainer_plan_detail', plan_id=plan.id)
        else:
            messages.error(request, "Error al crear el plan. Revisa los datos ingresados.")
    else:
        form = TrainingPlanForm()
        # Limitar clientes disponibles: Solo los asignados al user actual para privacidad.
        form.fields['client'].queryset = User.objects.filter(role='CLIENTE', assigned_professional=request.user)

    return render(request, 'entrenador/crear_plan.html', {'form': form})


# Detalle de plan para entrenador. Calcula progreso basado en logs (cualquier log cuenta como completado para flexibilidad).
#
# Por qué: Proporciona insights como progreso porcentual. Usamos round para 2 decimales en progress. Ordenamos workouts por semana/día para lógica temporal.

@login_required
def trainer_plan_detail(request, plan_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    plan = get_object_or_404(TrainingPlan, id=plan_id, trainer=request.user)
    workouts = plan.workouts.all().order_by('week_number', 'day_of_week')  # Orden lógico.

    total_exercises = WorkoutExercise.objects.filter(workout__plan=plan).count()
    completed_exercises = ExerciseLog.objects.filter(
        workout_exercise__workout__plan=plan
    ).count()  # Cuenta cualquier log; ajustable si se quiere solo 'completed'.
    progress = round((completed_exercises / total_exercises * 100), 2) if total_exercises > 0 else 0  # Evitar división por cero.
    context = {
        'plan': plan,
        'workouts': workouts,
        'progress': progress,
        'total_exercises': total_exercises,
        'completed_exercises': completed_exercises,
    }
    return render(request, 'entrenador/plan_detail.html', context)


# Edición de plan. Usa form con instance para actualizar.
#
# Por qué: Permite modificaciones simples. Mensajes para feedback.

@login_required
def edit_plan(request, plan_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    plan = get_object_or_404(TrainingPlan, id=plan_id, trainer=request.user)
    if request.method == 'POST':
        form = TrainingPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "Plan actualizado exitosamente.")
            return redirect('trainer_plan_detail', plan_id=plan.id)
    else:
        form = TrainingPlanForm(instance=plan)
    return render(request, 'entrenador/edit_plan.html', {'form': form, 'plan': plan})


#-------------------------------------------------------------------------------------------------------------------


#------------------------Gestión de Ejercicios y Workouts - Detalles Granulares--------------------------------

# Edición de un ejercicio en workout. Verifica ownership.
#
# Por qué: CRUD granular. Redirige a detail tras éxito para continuar editando.

@login_required
def edit_workout_exercise(request, exercise_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    w_exercise = get_object_or_404(WorkoutExercise, id=exercise_id, workout__plan__trainer=request.user)
    if request.method == 'POST':
        form = WorkoutExerciseForm(request.POST, instance=w_exercise)
        if form.is_valid():
            form.save()
            messages.success(request, "Ejercicio actualizado exitosamente.")
            return redirect('workout_detail', workout_id=w_exercise.workout.id)
    else:
        form = WorkoutExerciseForm(instance=w_exercise)
    return render(request, 'entrenador/edit_ejercicio.html', {'form': form, 'w_exercise': w_exercise})


# Agregar workout a un plan.
#
# Por qué: Parte de CRUD. Asigna plan automáticamente.

@login_required
def add_workout(request, plan_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    plan = get_object_or_404(TrainingPlan, id=plan_id, trainer=request.user)
    if request.method == 'POST':
        form = WorkoutForm(request.POST)
        if form.is_valid():
            workout = form.save(commit=False)
            workout.plan = plan
            workout.save()
            messages.success(request, "Entrenamiento agregado exitosamente.")
            return redirect('workout_detail', workout_id=workout.id)
    else:
        form = WorkoutForm()
    return render(request, 'entrenador/add_entrenamiento.html', {'form': form, 'plan': plan})


# Edición de workout.

@login_required
def edit_workout(request, workout_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    workout = get_object_or_404(Workout, id=workout_id, plan__trainer=request.user)
    if request.method == 'POST':
        form = WorkoutForm(request.POST, instance=workout)
        if form.is_valid():
            form.save()
            messages.success(request, "Entrenamiento actualizado exitosamente.")
            return redirect('workout_detail', workout_id=workout.id)
    else:
        form = WorkoutForm(instance=workout)
    return render(request, 'entrenador/edit_workout.html', {'form': form, 'workout': workout})


# Agregar ejercicio a workout.

@login_required
def add_exercise(request, workout_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    workout = get_object_or_404(Workout, id=workout_id, plan__trainer=request.user)
    if request.method == 'POST':
        form = WorkoutExerciseForm(request.POST)
        if form.is_valid():
            exercise = form.save(commit=False)
            exercise.workout = workout
            exercise.save()
            messages.success(request, "Ejercicio agregado exitosamente.")
            return redirect('workout_detail', workout_id=workout.id)  # Redirect back to detail to add more
    else:
        form = WorkoutExerciseForm()
    return render(request, 'entrenador/add_ejercicio.html', {'form': form, 'workout': workout})


# Eliminar ejercicio de workout.
#
# Por qué: Completa CRUD. Redirige a detail.

@login_required
def delete_workout_exercise(request, exercise_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    w_exercise = get_object_or_404(WorkoutExercise, id=exercise_id, workout__plan__trainer=request.user)
    workout_id = w_exercise.workout.id
    w_exercise.delete()
    messages.success(request, "Ejercicio eliminado exitosamente.")
    return redirect('workout_detail', workout_id=workout_id)


#-------------------------------------------------------------------------------------------------------------------

#------------------------Gestión de Calentamientos - Recursos para Clientes--------------------------------

# Actualizar/crear calentamiento. Opcional por ID.
#
# Por qué: Recursos reutilizables. Redirige a dashboard.

@login_required
def update_warmup(request, warmup_id=None):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    if warmup_id:
        warmup = get_object_or_404(Warmup, id=warmup_id)
    else:
        warmup = None
    if request.method == 'POST':
        form = WarmupForm(request.POST, instance=warmup)
        if form.is_valid():
            form.save()
            return redirect('trainer_dashboard')
    else:
        form = WarmupForm(instance=warmup)
    return render(request, 'entrenador/actualizar_calentamiento.html', {'form': form})


#-------------------------------Vistas para Clientes ---------------------------------

# Dashboard para clientes. Calcula métricas personalizadas como consistencia (basada en workouts completos).
#
# Por qué: Vista personalizada para motivación. Loop para completed_workouts es aceptable por bajo volumen.

@login_required
def client_dashboard(request):
    if request.user.role != 'CLIENTE':
        return redirect('inicio')

    plans = TrainingPlan.objects.filter(client=request.user)
    active_plans_count = plans.filter(status='active').count()
    weekly_sessions = Workout.objects.filter(
        plan__in=plans,
        date__range=[timezone.now().date(), timezone.now().date() + timedelta(days=7)]
    ).exclude(date__isnull=True).count()
    completed_exercises = ExerciseLog.objects.filter(
        client=request.user, status='completed'
    ).count()
    next_session = Workout.objects.filter(
        plan__in=plans,
        date__gte=timezone.now().date()
    ).exclude(date__isnull=True).order_by('date').first()
    upcoming_sessions = Workout.objects.filter(
        plan__in=plans,
        date__gte=timezone.now().date()
    ).exclude(date__isnull=True).order_by('date')[:5]
    total_workouts = Workout.objects.filter(plan__in=plans).count()
    total_exercises = WorkoutExercise.objects.filter(workout__plan__in=plans).count()
    completed_workouts = 0
    for workout in Workout.objects.filter(plan__in=plans):
        workout_exercises_count = workout.exercises.count()
        completed_logs = ExerciseLog.objects.filter(
            workout_exercise__workout=workout,
            status='completed'
        ).count()
        if workout_exercises_count > 0 and completed_logs == workout_exercises_count:
            completed_workouts += 1
    consistency = round((completed_workouts / total_workouts * 100), 2) if total_workouts > 0 else 0
    warmups = Warmup.objects.all()
    warmups_upper_body = warmups.filter(type='superior')
    warmups_lower_body = warmups.filter(type='inferior')
    context = {
        'plans': plans,
        'active_plans_count': active_plans_count,
        'weekly_sessions': weekly_sessions,
        'completed_exercises': completed_exercises,
        'next_session': next_session,
        'upcoming_sessions': upcoming_sessions,
        'total_workouts': total_workouts,
        'total_exercises': total_exercises,
        'consistency': consistency,
        'warmups': {
            'upper_body': warmups_upper_body,
            'lower_body': warmups_lower_body,
        }
    }
    return render(request, 'clientes/cliente.html', context)


# Estadísticas de cliente: Lista logs ordenados.

@login_required
def client_statistics(request):
    if request.user.role != 'CLIENTE':
        return redirect('inicio')
    logs = ExerciseLog.objects.filter(client=request.user).order_by('-date_completed')
    context = {'logs': logs}
    return render(request, 'clientes/estadisticas.html', context)


# Ver plan para cliente. Calcula progreso similar al trainer.

@login_required
def view_plan(request, plan_id):
    plan = get_object_or_404(TrainingPlan, id=plan_id, client=request.user)
    workouts = plan.workouts.all().order_by('week_number', 'day_of_week')
    total_exercises = WorkoutExercise.objects.filter(workout__plan=plan).count()
    completed_exercises = ExerciseLog.objects.filter(
        workout_exercise__workout__plan=plan
    ).count()
    progress = round((completed_exercises / total_exercises * 100), 2) if total_exercises > 0 else 0
    completed_workouts = 0
    for workout in workouts:
        workout_exercises = workout.exercises.count()
        completed_workout_exercises = ExerciseLog.objects.filter(
            workout_exercise__workout=workout, status='completed'
        ).count()
        if workout_exercises > 0 and completed_workout_exercises == workout_exercises:
            completed_workouts += 1
    context = {
        'plan': plan,
        'workouts': workouts,
        'progress': progress,
        'total_exercises': total_exercises,
        'completed_exercises': completed_exercises,
        'completed_workouts': completed_workouts,
    }
    return render(request, 'clientes/ver_plan.html', context)


# Registrar log de ejercicio. Envía email al trainer con detalles.
@login_required
def log_exercise(request, workout_exercise_id):
    workout_exercise = get_object_or_404(WorkoutExercise, id=workout_exercise_id, workout__plan__client=request.user)
    best_log = ExerciseLog.objects.filter(
        client=request.user,
        workout_exercise__exercise=workout_exercise.exercise
    ).order_by('-weight_lifted_kg', '-reps_completed').first()
    
    if request.method == 'POST':
        form = ExerciseLogForm(request.POST, request.FILES)
        if form.is_valid():
            if workout_exercise.video_required and ('video_log' not in request.FILES and 'video_key' not in request.POST):
                messages.error(request, "Este ejercicio requiere un video.")
                return render(request, 'clientes/report.html', {'form': form, 'workout_exercise': workout_exercise, 'best_log': best_log})
            
            log = form.save(commit=False)
            log.client = request.user
            log.workout_exercise = workout_exercise
            
            if 'video_key' in request.POST:
                log.video_log.name = request.POST['video_key']
            elif 'video_log' in request.FILES:
                video_file = request.FILES['video_log']
                mime_type, _ = mimetypes.guess_type(video_file.name)
                if not mime_type or not mime_type.startswith('video/'):
                    messages.error(request, "El archivo subido no es un video válido.")
                    return render(request, 'clientes/report.html', {'form': form, 'workout_exercise': workout_exercise, 'best_log': best_log})
            
            log.save()  # Sube a B2 si aplica (pero ahora es direct desde client)
            
            workout = workout_exercise.workout
            if workout.is_complete():
                send_daily_report(workout)
            
            return redirect('view_plan', plan_id=workout_exercise.workout.plan.id)
    else:
        form = ExerciseLogForm()
    
    context = {
        'form': form,
        'workout_exercise': workout_exercise,
        'best_log': best_log,
    }
    return render(request, 'clientes/report.html', context)

# View para presigned simple (mantenido para fallbacks, pero usaremos multipart)
@require_POST
@login_required
def generate_presigned_url(request):
    file_name = request.POST.get('file_name')
    if not file_name:
        return JsonResponse({'error': 'No file name provided'}, status=400)
    
    import os
    ext = os.path.splitext(file_name)[1][1:].lower()
    if ext not in ['mp4', 'avi', 'mov']:
        return JsonResponse({'error': 'Extension not allowed'}, status=400)
    
    mime_type, _ = mimetypes.guess_type(file_name)
    if not mime_type or not mime_type.startswith('video/'):
        return JsonResponse({'error': 'Invalid file type'}, status=400)
    
    unique_key = f'logs/videos/{uuid.uuid4()}_{file_name}'
    
    s3_client = boto3.client(
        's3',
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    
    try:
        presigned_url = s3_client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                'Key': unique_key,
                'ContentType': mime_type
            },
            ExpiresIn=7200  # Aumentado a 2 horas
        )
        return JsonResponse({'url': presigned_url, 'key': unique_key})
    except ClientError as e:
        return JsonResponse({'error': str(e)}, status=500)

# Nuevos views para multipart
@require_POST
@login_required
def initiate_multipart_upload(request):
    file_name = request.POST.get('file_name')
    if not file_name:
        return JsonResponse({'error': 'No file name provided'}, status=400)
    
    import os
    ext = os.path.splitext(file_name)[1][1:].lower()
    if ext not in ['mp4', 'avi', 'mov']:
        return JsonResponse({'error': 'Extension not allowed'}, status=400)
    
    mime_type, _ = mimetypes.guess_type(file_name)
    if not mime_type or not mime_type.startswith('video/'):
        return JsonResponse({'error': 'Invalid file type'}, status=400)
    
    unique_key = f'logs/videos/{uuid.uuid4()}_{file_name}'
    
    s3_client = boto3.client(
        's3',
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    
    try:
        multipart = s3_client.create_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=unique_key,
            ContentType=mime_type
        )
        return JsonResponse({'upload_id': multipart['UploadId'], 'key': unique_key})
    except ClientError as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_POST
@login_required
def generate_presigned_part(request):
    key = request.POST.get('key')
    upload_id = request.POST.get('upload_id')
    part_number = int(request.POST.get('part_number'))
    
    s3_client = boto3.client(
        's3',
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    
    try:
        url = s3_client.generate_presigned_url(
            'upload_part',
            Params={
                'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                'Key': key,
                'UploadId': upload_id,
                'PartNumber': part_number
            },
            ExpiresIn=7200  # 2 horas
        )
        return JsonResponse({'url': url})
    except ClientError as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_POST
@login_required
def complete_multipart_upload(request):
    key = request.POST.get('key')
    upload_id = request.POST.get('upload_id')
    parts = json.loads(request.POST.get('parts'))  # Lista de {'ETag': etag, 'PartNumber': num}
    
    s3_client = boto3.client(
        's3',
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY
    )
    
    try:
        s3_client.complete_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            UploadId=upload_id,
            MultipartUpload={'Parts': parts}
        )
        return JsonResponse({'success': True})
    except ClientError as e:
        return JsonResponse({'error': str(e)}, status=500)

def send_daily_report(workout):
    # Generar Excel con diseño mejorado
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte Diario - {workout.title}"
    
    # Agregar título principal
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Reporte Diario Completado: {workout.title} por {workout.plan.client.username}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar fila en blanco para separación
    ws.append([])  # Fila 2 vacía
    
    # Encabezados
    headers = ['Ejercicio', 'Peso (kg)', 'Reps', 'RIR', 'RPE', 'Estado', 'Notas', 'Video URL']
    ws.append(headers)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="17375E", end_color="17375E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Obtener logs
    logs = ExerciseLog.objects.filter(workout_exercise__workout=workout)
    
    # Colores para estados (para resaltar filas según estado)
    status_colors = {
        'completed': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Verde claro
        'half': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),      # Amarillo claro
        'not_completed': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rojo claro
    }
    
    for log in logs:
        video_url = log.video_log.url if log.video_log else ''
        row_data = [
            log.workout_exercise.exercise.name,
            log.weight_lifted_kg,
            log.reps_completed,
            log.rir_actual if log.rir_actual is not None else '',
            log.rpe_actual if log.rpe_actual is not None else '',
            log.get_status_display(),
            log.notes,
            video_url
        ]
        ws.append(row_data)
        
        # Obtener número de fila actual
        current_row = ws.max_row
        
        # Aplicar borde y alineación a las celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment
            
            # Formatear números
            if col in [2, 3, 4, 5]:  # Peso, Reps, RIR, RPE
                cell.number_format = '0.00' if col == 2 else '0'
            
            # Hacer URL de video clickable si existe
            if col == 8 and video_url:
                cell.hyperlink = video_url
                cell.font = Font(underline="single", color="0000FF")
                cell.value = "Ver Video" if video_url else ''
        
        # Resaltar fila según estado
        status_fill = status_colors.get(log.status, None)
        if status_fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = status_fill
    
    # Ajustar ancho de columnas automáticamente
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar la fila de encabezados
    ws.freeze_panes = 'A4'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Mejorar el email: usar HTML para un cuerpo más atractivo
    trainer_email = workout.plan.trainer.email
    subject = f"Reporte Diario Completado: {workout.title} por {workout.plan.client.username}"
    
    html_message = f"""
    <html>
        <body>
            <h2>Reporte Diario Completado</h2>
            <p>Estimado entrenador,</p>
            <p>El cliente {workout.plan.client.username} ha completado el workout "{workout.title}" el {workout.date.strftime('%Y-%m-%d')}.</p>
            <p>Adjunto encontrarás el reporte detallado en formato Excel, con los logs de ejercicios, incluyendo pesos, reps, RIR/RPE, estados, notas y enlaces a videos si están disponibles.</p>
            <p>Gracias,</p>
            <p>Equipo de la App</p>
        </body>
    </html>
    """
    
    email = EmailMessage(subject, html_message, settings.EMAIL_HOST_USER, [trainer_email])
    email.content_subtype = "html"  # Para que se envíe como HTML
    email.attach(f'reporte_diario_{workout.date.strftime("%Y-%m-%d")}.xlsx', output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()


@login_required
def progress_view(request, plan_id):
    plan = get_object_or_404(TrainingPlan, id=plan_id)
    if request.user.role == 'ENTRENADOR' and plan.trainer != request.user:
        messages.error(request, "No tienes permiso.")
        return redirect('inicio')
    elif request.user.role == 'CLIENTE' and plan.client != request.user:
        messages.error(request, "No tienes permiso.")
        return redirect('inicio')
    # Datos para charts: Por ejercicio, fechas y pesos/reps
    exercises = Exercise.objects.filter(workoutexercise__workout__plan=plan).distinct()
    chart_data = {}
    for ex in exercises:
        logs = ExerciseLog.objects.filter(workout_exercise__exercise=ex, client=plan.client).order_by('date_completed')
        dates = [log.date_completed.strftime('%Y-%m-%d') for log in logs]
        weights = [log.weight_lifted_kg for log in logs]
        reps = [log.reps_completed for log in logs]
        chart_data[ex.name] = {
            'dates': dates,
            'weights': weights,
            'reps': reps
        }
    context = {'plan': plan, 'chart_data': chart_data}
    return render(request, 'entrenador/progress.html', context)

# ---------------------------------------------------------------------------------------------------------------
#------------------------Vistas Adicionales - Logs y Nutricionistas --------------------------------
# Ver detalle de log. Verifica permisos basados en rol.
@login_required
def view_log(request, log_id):
    log = get_object_or_404(ExerciseLog, id=log_id)
    if request.user.role == 'ENTRENADOR' and log.workout_exercise.workout.plan.trainer == request.user:
        pass
    elif request.user.role == 'CLIENTE' and log.client == request.user:
        pass
    else:
        messages.error(request, "No tienes permiso para ver este registro.")
        return redirect('inicio')
    context = {'log': log}
    return render(request, 'entrenador/view_log.html', context)





# Logs por cliente para entrenador.

@login_required
def client_logs(request, client_id):
    if request.user.role != 'ENTRENADOR':
        return redirect('inicio')
    client = get_object_or_404(User, id=client_id, assigned_professional=request.user)
    logs = ExerciseLog.objects.filter(client=client).order_by('-date_completed')
    context = {
        'client': client,
        'logs': logs,
    }
    return render(request, 'entrenador/client_logs.html', context)





@login_required
def create_exercise(request):
    if request.user.role != 'ENTRENADOR':
        messages.error(request, "No tienes permiso para crear ejercicios.")
        return redirect('inicio')

    if request.method == 'POST':
        form = ExerciseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ejercicio creado exitosamente.")
            return redirect('trainer_dashboard')
        else:
            messages.error(request, "Error al crear el ejercicio. Revisa los datos ingresados.")
    else:
        form = ExerciseForm()

    return render(request, 'entrenador/create_exercise.html', {'form': form})



# Dashboard para nutricionistas: Placeholder para futura implementación.

@login_required
def nutritionist_dashboard(request):
    if request.user.role != 'NUTRICIONISTA':
        return redirect('inicio')
    # Implement as needed
    return render(request, 'nutricionistas/nutricionistas.html')


